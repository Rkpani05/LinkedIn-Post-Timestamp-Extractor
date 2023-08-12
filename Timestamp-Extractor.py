import tkinter as tk
from tkinter import ttk, filedialog
import re
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook

# Function for the Single URL Timestamp Extractor
def show_single_extractor():

    def get_post_id(url):
        match = re.search(r'([0-9]{19})', url)
        if match:
            return match.group(1)
        return None

    def extract_unix_timestamp(post_id):
        as_binary = bin(int(post_id))[2:]
        first41_chars = as_binary[:41]
        timestamp = int(first41_chars, 2)
        return timestamp

    def unix_timestamp_to_human_date(timestamp, to_ist):
        timestamp = timestamp / 1000
        utc_time = datetime.fromtimestamp(timestamp, tz=timezone.utc)
        if to_ist:
            ist_time = utc_time + timedelta(hours=5, minutes=30)
            return ist_time.strftime('%a, %d %b %Y %I:%M:%S %p IST')
        else:
            return utc_time.strftime('%a, %d %b %Y %I:%M:%S %p UTC')

    def get_date():
        url = url_entry.get()
        post_id = get_post_id(url)
        if post_id:
            unix_timestamp = extract_unix_timestamp(post_id)
            to_ist = timezone_var.get() == "IST"
            human_date = unix_timestamp_to_human_date(unix_timestamp, to_ist)
            date_label.config(text=human_date)
        else:
            date_label.config(text="Invalid URL or Post ID not found.")

    single_extractor = tk.Toplevel(app)
    single_extractor.title("LinkedIn Single URL Timestamp Extractor")
    single_extractor.state('zoomed')
    single_extractor.configure(bg='azure')

    heading = ttk.Label(single_extractor, text="Single URL TimeStamp Extractor", background='azure', font=("Arial", 16, "bold"))
    heading.pack(pady=20)

    frame = ttk.Frame(single_extractor, padding="20")
    frame.pack(pady=40, expand=True)

    url_label = ttk.Label(frame, text="Enter LinkedIn URL:")
    url_label.grid(column=0, row=0, sticky=tk.W)

    url_entry = ttk.Entry(frame, width=50)
    url_entry.grid(column=1, row=0, sticky=(tk.W, tk.E), padx=10)

    timezone_var = tk.StringVar()
    timezone_var.set("UTC")
    timezone_label = ttk.Label(frame, text="Timezone:")
    timezone_label.grid(column=0, row=1, sticky=tk.W)

    timezone_dropdown = ttk.Combobox(frame, textvariable=timezone_var, values=["UTC", "IST"], width=28)  
    timezone_dropdown.grid(column=1, row=1, sticky=(tk.W, tk.E), padx=10)

    get_date_btn = ttk.Button(frame, text="Submit", command=get_date)
    get_date_btn.grid(column=1, row=2, pady=20)

    timestamp_label = ttk.Label(frame, text="Post Timestamp:")
    timestamp_label.grid(column=0, row=3, sticky=tk.W)

    date_label = ttk.Label(frame, text="", foreground= "deep sky blue")
    date_label.grid(column=1, row=3, pady=5)

    close_btn = ttk.Button(frame, text="Close", command=single_extractor.destroy)
    close_btn.grid(column=1, row=4, pady=20)

# Function for the Multiple URLs Timestamp Extractor
def show_multiple_extractor():

    def get_post_id(url):
        if not isinstance(url, str):  
            return None
        match = re.search(r'([0-9]{19})', url)
        if match:
            return match.group(1)
        return None

    def extract_unix_timestamp(post_id):
        as_binary = bin(int(post_id))[2:]
        first41_chars = as_binary[:41]
        timestamp = int(first41_chars, 2)
        return timestamp

    def unix_timestamp_to_human_date(timestamp, to_ist):
        timestamp = timestamp / 1000
        utc_time = datetime.fromtimestamp(timestamp, tz=timezone.utc)
        if to_ist:
            ist_time = utc_time + timedelta(hours=5, minutes=30)
            return ist_time.strftime('%a, %d %b %Y %I:%M:%S %p IST')
        else:
            return utc_time.strftime('%a, %d %b %Y %I:%M:%S %p UTC')

    def process_xlsx_file():
        input_file = filedialog.askopenfilename(title="Select Input Excel", filetypes=[("Excel Files", "*.xlsx")])
        if not input_file:
            return

        workbook = load_workbook(input_file)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        if not 'Timestamp UTC' in headers and not 'Timestamp IST' in headers:
            sheet.cell(row=1, column=3, value="Timestamp UTC")
            sheet.cell(row=1, column=4, value="Timestamp IST")
            start_row = 2
        else:
            start_row = 2
            while sheet.cell(row=start_row, column=3).value and sheet.cell(row=start_row, column=4).value:
                start_row += 1

        for idx, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, values_only=True), start=start_row):
            if len(row) >= 2:  # At least ID and URL present
                id_val, url = row[:2]
                post_id = get_post_id(url)
                if post_id:
                    unix_timestamp = extract_unix_timestamp(post_id)
                    utc_time = unix_timestamp_to_human_date(unix_timestamp, to_ist=False)
                    ist_time = unix_timestamp_to_human_date(unix_timestamp, to_ist=True)
                    
                    sheet.cell(row=idx, column=3, value=utc_time)
                    sheet.cell(row=idx, column=4, value=ist_time)

        workbook.save(input_file)

        for r in results_tree.get_children():
            results_tree.delete(r)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            results_tree.insert("", "end", values=row)
        status_label.config(text="Processing Complete!")

    multiple_extractor = tk.Toplevel(app)
    multiple_extractor.title("LinkedIn Multiple URL Timestamp Extractor")
    multiple_extractor.state('zoomed')
    multiple_extractor.configure(bg='azure')

    heading = ttk.Label(multiple_extractor, text="Multiple URL TimeStamp Extractor", background='azure', font=("Arial", 16, "bold"))
    heading.pack(pady=20)

    frame = ttk.Frame(multiple_extractor, padding="20")
    frame.pack(pady=40, expand=True)

    import_label = ttk.Label(frame, text="Import Excel File:", background='azure', font=("Arial", 12))
    import_label.grid(column=0, row=0, sticky=tk.W, pady=20)

    process_btn = ttk.Button(frame, text="Import & Process Excel", command=process_xlsx_file)
    process_btn.grid(column=1, row=0, pady=20)

    results_tree = ttk.Treeview(frame, columns=("ID", "URL", "Timestamp UTC", "Timestamp IST"), show="headings")
    results_tree.heading("ID", text="ID")
    results_tree.heading("URL", text="URL")
    results_tree.heading("Timestamp UTC", text="Timestamp UTC")
    results_tree.heading("Timestamp IST", text="Timestamp IST")
    results_tree.grid(column=0, row=1, columnspan=2, sticky=(tk.W, tk.E), padx=10, pady=10)

    scrollbar = ttk.Scrollbar(frame, orient="vertical", command=results_tree.yview)
    scrollbar.grid(column=2, row=1, sticky="ns")
    results_tree.configure(yscrollcommand=scrollbar.set)

    status_label = ttk.Label(frame, text="", foreground="green")
    status_label.grid(column=1, row=2, pady=5)
    note_label = ttk.Label(frame, text=(
    "NOTE: Ensure the Excel file (.xlsx) contains at least two columns:\n"
    "1. ID - Unique identifier for each row.\n"
    "2. URL - The LinkedIn URL containing the post ID.\n"
    "The program will process the file and add the 'Timestamp UTC' and 'Timestamp IST' columns."
    ), background='azure', wraplength=1000)  
    note_label.grid(column=0, row=3, columnspan=3, pady=20)

    close_btn = ttk.Button(frame, text="Close", command=multiple_extractor.destroy)
    close_btn.grid(column=1, row=4, pady=20)  


app = tk.Tk()
app.title("LinkedIn Post Timestamp Extractor")
app.state('zoomed')
app.configure(bg='azure')

style = ttk.Style()
style.configure("TLabel", font=("Arial", 14, "bold"), padding=10)
style.configure("TButton", font=("Arial", 12, "bold"), padding=10)

heading_label = ttk.Label(app, text="LinkedIn Post Timestamp Extractor", background='azure')
heading_label.pack(pady=20)

middle_x = 0.5
middle_y = 0.45  # Slightly above the middle to account for two buttons

single_btn = ttk.Button(app, text="Single Link", command=show_single_extractor)
single_btn.place(relx=middle_x, rely=middle_y, anchor='center')

multiple_btn = ttk.Button(app, text="Multiple Links", command=show_multiple_extractor)
multiple_btn.place(relx=middle_x, rely=middle_y + 0.1, anchor='center')  # 0.05 to move it slightly below the first button

app.mainloop()
